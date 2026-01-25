import openpyxl
import json


def parse_schedule_correctly(file_path):
    """Парсит школьное расписание правильно"""

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Печатаем информацию о таблице для отладки
    print(f"Всего строк: {ws.max_row}, Всего колонок: {ws.max_column}")

    # Ищем строку с классами - сканируем первые 10 строк
    class_row = None
    class_columns = {}

    for row in range(1, 10):
        # Ищем строку, где есть '7а' или '10б' и т.д.
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and (
                    '7а' in str(cell_value) or '8а' in str(cell_value) or '9а' in str(cell_value) or '10а' in str(
                    cell_value) or '11а' in str(cell_value)):
                class_row = row
                print(f"Найдена строка с классами: строка {row}")
                break
        if class_row:
            break

    if not class_row:
        print("Не удалось найти строку с классами!")
        return {}

    # Собираем названия классов
    classes = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=class_row, column=col).value
        if cell_value:
            class_name = str(cell_value).strip()
            # Проверяем, что это действительно название класса
            if any(x in class_name for x in ['7а', '7б', '8а', '8б', '9а', '9б', '10а', '10б', '11а', '11б']):
                classes.append(class_name)
                class_columns[col] = class_name
                print(f"Класс {class_name} в колонке {col}")

    print(f"\nНайдено классов: {len(classes)}")
    print(f"Классы: {classes}")

    # Создаем структуру данных
    schedule = {class_name: {} for class_name in classes}

    current_day = None

    # Проходим по всем строкам ниже строки с классами
    for row in range(class_row + 1, ws.max_row + 1):
        # День недели в колонке A
        day_value = ws.cell(row=row, column=1).value
        if day_value and str(day_value).strip():
            day_str = str(day_value).strip()
            if day_str in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                current_day = day_str
                # print(f"День: {current_day}")

        # Номер урока в колонке B
        lesson_value = ws.cell(row=row, column=2).value
        if lesson_value is None:
            continue

        # Преобразуем номер урока
        try:
            lesson_num = int(float(lesson_value))
        except:
            # Пропускаем нечисловые значения
            continue

        # Если нет текущего дня, пропускаем
        if not current_day:
            continue

        # Для каждого класса получаем предмет
        for col, class_name in class_columns.items():
            # Инициализируем день для класса
            if current_day not in schedule[class_name]:
                schedule[class_name][current_day] = {}

            # Получаем предмет
            subject = ws.cell(row=row, column=col).value

            if subject and str(subject).strip():
                subject_name = str(subject).strip()
                # Проверяем, не занят ли уже этот урок
                if lesson_num not in schedule[class_name][current_day]:
                    schedule[class_name][current_day][lesson_num] = subject_name

    return schedule


def print_schedule_for_all_classes(schedule):
    """Печатает расписание для всех классов"""

    days_order = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']

    for class_name in sorted(schedule.keys()):
        print(f"\n{'=' * 60}")
        print(f"РАСПИСАНИЕ КЛАССА: {class_name}")
        print(f"{'=' * 60}")

        class_data = schedule[class_name]

        for day in days_order:
            if day in class_data:
                print(f"\n{day}:")
                print("-" * 40)

                day_lessons = class_data[day]
                for lesson_num in sorted(day_lessons.keys()):
                    print(f"  {lesson_num}. {day_lessons[lesson_num]}")


def save_schedule_to_json(schedule, filename="school_schedule.json"):
    """Сохраняет расписание в JSON файл"""
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(schedule, f, ensure_ascii=False, indent=2)
    print(f"\nРасписание сохранено в {filename}")


# Запуск парсера
if __name__ == "__main__":
    print("=== ПАРСИНГ ШКОЛЬНОГО РАСПИСАНИЯ ===")

    file_path = "schedule.xlsx"

    try:
        schedule = parse_schedule_correctly(file_path)

        if schedule:
            print_schedule_for_all_classes(schedule)
            save_schedule_to_json(schedule)
        else:
            print("Не удалось распарсить расписание")

    except FileNotFoundError:
        print(f"Файл {file_path} не найден!")
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback

        traceback.print_exc()